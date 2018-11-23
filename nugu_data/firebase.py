import os, time, re
import json
from dateutil.parser import parse
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db


class Firebase():


    def __init__(self):
        # Initialize firebase.
        _token = {
            "type": "service_account",
            "project_id": "foodreet-e783e",
            "private_key_id": "3d49c38eede5091cd953e258492488b12611fad1",
            "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDbd0jNlUE1adew\nul8E9WiIFrpa14mal4wiNxqMyLYfYBrTW1pGuc48L9J9HlsXk/doSV/9xEHNuS1a\nU25+Ou4RIRpxrEpwrWuipmVOQsgnmQpLrvKYkNk2O0ovRgPuzhboq24qO8wX8sUD\nwe7g7YEiTIdcXjSYkr5cCAfGn5sBpxcYhOEvg4fzNbeSZeZOX7D7mSnSHXjgnJ9o\n/EItQ9pe7BIkYaDD4RgQBxNHQEyjtvkVN8W2wmlbAlpW8ydbVCHRN2NW/Job0T65\nyoqK4Mg3VDWRN/WtuQgj6k3/y0ckfkTJW50CZ3u2MJHwhjDrjvhJ++g6UCvq6uVB\nTMr1LTI/AgMBAAECggEAAobQiUFfI1DJIWk739A9eFeepJ4XoPZ+sAQnY5iJ9DmE\nhYtnMipAdBPu3e7kCJT1A7fYiZDhWw93NLOCMaTXCc5yQf3ES6DLfO9iejCtojsS\nUcW9WEFOpt+u3nV1ZSHtRJd4qLbQhjQOzZjFecRyOKVd55PUEi+aTm/vOvuqaq3u\nz+KcVfg3ct8EYGnGMEx9itDgWzFnlBDAvzSykEVDCJDFoLH1hl4S5NNx5tWCGiJR\nZPafZUWqSJTf1xnNRlafqY/1K6rJ6D40bRuARMdlkdMUFgUFMnJUDelGu/lQmQ9/\nUlB/OViDIJGNcs7M47/2hiP9m4ldrATFeGbkxJjtoQKBgQD72U1Zp+nCVFo/gJf/\nCJxoQgluPdJXMOYbqKsjjflzOaU5DRqMVfvVVdpbdTgKqQP1IP0szetg+jSFqo7J\nGAR54utD0194SXy+vej3FFa2QrKvq7UKdQ0NsBJY5ab91lgUgWwJxatGCuEuqkBC\nFOckkwK2fkyoNRWpp9TWu7rA4QKBgQDfFVcCVAJ5FEQO6TUD6xwQd5h+/bn9ZEbH\nn+7NM592pGjhwREOjOhlMzBfgiiDPuzbJJKxX+ms2o7q5+P8f2a6XAT2E9Q+JGbv\nKuYHaGcBFdT0+7MGi8qDyWduUOvwSC3c4K5qvpsjJmoEUY4i51TGndxFEK1uixnG\nUDn5fI+3HwKBgQDVXNm9FRZjPlde6TDcLY1kfl3C+bkTiSfJUvH6IzOsPLAPT1yo\ne8foit19EieyUFI8AAC9To8VD6QjgcWJ9EEGigpK4e5F1+xJJFcTzDVSH9uGHHA6\nNcQjToGcsqLw9gCVDrha5w4dPuZbRH0//rl8MNfhh8qlzqNVogwdB9pFwQKBgQDC\nfmZ3nyrwyVx065nPKsIzETIhNbFcXHrhPG0E9wcGd10BxkPKvcs3wCNigaY0lz4F\n6aMDhi3DCGImJ7q7v4+e1u90Y+FdtHsR9gLLXpi/d3tNg8yX3nTX5VaJ9X9e5G8n\n7bPjWP8E7E7uKPSVNb7R03vhsyKiCCCWqAeDnrfl3QKBgGRukLkGNbrZXOWFgYhP\n2/Zr0NWWK9uwZthESMvVuMWsF2ZdeP9iYCU+8546LKMQPmj4aF7phZ4qPyE5l/1b\nV8nA8p1DD9Jq0HDYPBOl/VP7A7aPy2qrU8PvDqUMqX46isyXbEKJh8yNo76n+WYa\nDkX7tZHzOVMSr7RovQEPTqWq\n-----END PRIVATE KEY-----\n",
            "client_email": "firebase-adminsdk-6vdgn@foodreet-e783e.iam.gserviceaccount.com",
            "client_id": "105473880117782672746",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-6vdgn%40foodreet-e783e.iam.gserviceaccount.com"
        }
        cred = credentials.Certificate(json.loads(str(_token).replace("'", '"')))
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred, {
                "databaseURL": "https://foodreet-e783e.firebaseio.com/"
            })
        self.db = db.reference('calories') # DB Reference
        self.bmi = {} # BMI Reference
        self.ret = {} # Temporary Calorie Result
        self.user_info = {} # User Infomation Reference
        self.file_dir = "/Users/moonseongjae/python-worksapce/Nugu/nugu_data/calorie/calorie{}.xlsx"
        self.fm_pa_list = [1.0, 1.12, 1.27, 1.45] # 여성 신체활동계수
        self.m_pa_list = [1.0, 1.11, 1.25, 1.48] # 남성 신체활동계수
        # License, Arizona State University (https://www.ncbi.nlm.nih.gov/pubmed/21681120)
        # 2011 Compendium of Physical Activities: a second update of codes and MET values.
        self.exercise = [
            ["자전거", 7.5], # 기본
            ["걷기", 8.3], # 기본
            ["조깅", 7], # 기본
            ["수영", 7], # 자유형 기준
            ["야구", 6.5], # 야구 기본
            ["볼링", 3.8], # 볼링 기본
            ["복싱", 12.8], # In Ring
            ["축구", 7], # 축구 기본
            ["스쿼시", 7.3], # 스쿼시 기본
            ["테니스", 7.3, 0], # 테니스 기본
            ["배구", 7, 0], # 배구 기본
            ["코어", 3.8, 0], # 코어 운동 기본
            ["필라테스", 3, 0], # 필라테스 기본
            ["PT", 7.8, 0], # PT 기본
            ["에어로빅", 7.3, 0], # 에어로빅 기본
            ["헬스", 5.5, 0] # 헬스 기본
        ]


    def save_cal2(self, file):
        if not file:
            raise Exception('NoFileName')
        if not os.path.exists(self.file_dir.format(file)):
            raise Exception('NoSuchFileName')
        try:
            wb = load_workbook(self.file_dir.format(file), read_only=True)
            sheet = wb['calorie']
            for r in sheet.rows:
                category = r[1].value
                name = r[3].value
                quanty = r[4].value
                kal = r[5].value
                tan_g = r[6].value
                dan_g = r[7].value
                ji_g = r[8].value
                dang_g = r[9].value
                na_mg = r[10].value
                chol_mg = r[11].value
                pho_g = r[12].value
                trans_g = r[13].value
                data = {
                    "category": category,
                    "name": name,
                    "quanty": quanty,
                    "kal": kal,
                    "tan_g": tan_g,
                    "dan_g": dan_g,
                    "ji_g": ji_g,
                    "dang_g": dang_g,
                    "na_mg": na_mg,
                    "chol_mg": chol_mg,
                    "pho_g": pho_g,
                    "trans_g": trans_g
                }
                self.db.child("calories").push(data)
            print('--- save done ---')
        except Exception as e:
            print(e)
        finally:
            wb.close()


    def save_cal13(self, file=''):
        if not os.path.exists(self.file_dir.format(FILE)):
            raise Exception('NoSuchFileName')
        try:
            wb = load_workbook(self.file_dir.format(FILE), read_only=True)
            sheet = wb['calorie']
            for r in sheet.rows:
                category = r[1].value
                name = r[2].value
                quanty = r[3].value
                kal = r[4].value
                tan_g = r[5].value
                dan_g = r[6].value
                ji_g = r[7].value
                dang_g = r[8].value
                na_mg = r[9].value
                chol_mg = r[10].value
                pho_g = r[11].value
                trans_g = r[12].value
                data = {
                    "category": category,
                    "name": name,
                    "quanty": quanty,
                    "kal": kal,
                    "tan_g": tan_g,
                    "dan_g": dan_g,
                    "ji_g": ji_g,
                    "dang_g": dang_g,
                    "na_mg": na_mg,
                    "chol_mg": chol_mg,
                    "pho_g": pho_g,
                    "trans_g": trans_g
                }
                db.child("calories").push(data, self.user['idToken'])
        except Exception as e:
            print(e)
        finally:
            wb.close()


    def edit_distance(self, s1, s2):
        l1, l2 = len(s1), len(s2)
        if l2 > l1:
            return self.edit_distance(s2, s1)
        if l2 is 0:
            return l1
        prev_row = list(range(l2 + 1))
        current_row = [0] * (l2 + 1)
        for i, c1 in enumerate(s1):
            current_row[0] = i + 1
            for j, c2 in enumerate(s2):
                d_ins = current_row[j] + 1
                d_del = prev_row[j + 1] + 1
                d_sub = prev_row[j] + (1 if c1 != c2 else 0)
                current_row[j + 1] = min(d_ins, d_del, d_sub)
            prev_row[:] = current_row[:]
        return prev_row[-1]


    def set_cal(self, name):
        '''
        Usage
            Get calroie list from Firebase and initialize self.ret.
        Param
            name: User Speech(Entities).
        '''
        try:
            save_distance = [] # temp var
            data = self.db.get()
            if data is not None:
                # Loop for entities list.
                for d in data.values():
                    if name in d.get('name'):
                        temp = {
                            "category": d.get('category'),
                            "name": d.get('name'), 
                            "quanty": d.get('quanty'),
                            "kal": round(d.get('kal')),
                            "tan_g": d.get('tan_g'),
                            "dan_g": d.get('dan_g'),
                            "ji_g": d.get('ji_g'),
                            "dang_g": d.get('dang_g'),
                            "na_mg": d.get('na_mg'),
                            "chol_mg": d.get('chol_mg'),
                            "pho_g": d.get('pho_g'),
                            "trans_g": d.get('trans_g')   
                        }
                        save_distance.append({'cal':temp})
                # Calculate for edit_distance.
                if save_distance and save_distance is not None:
                    distance = min(save_distance, key = lambda s: self.edit_distance(name, s.get("cal").get("name")))
                    if distance and distance is not None:
                        na_mg = distance.get("cal").get("na_mg")
                        if na_mg > 1000:
                            distance["content"] = "이 음식은 " + str(round(na_mg)) + "mg의 높은 나트륨 수치의 음식입니다."
                    self.ret['result'] = distance
                else:
                    self.ret['result'] = None
        except Exception as e:
            print(e)


    def get_percent(self):
        
        _json = json.loads(str(self.ret).replace("\'", '"'))

        kal_sum, quanty_sum, tan_g_sum, dan_g_sum, ji_g_sum, \
        dang_g_sum, na_mg_sum, chol_mg_sum, pho_g_sum, trans_g_sum = \
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0
        
        # cn_list = []
        for t in _json.get('result'):
            # cn_list.append(t.get('cal').get('name'))
            # cn_list.append(t.get('cal').get('category'))
            kal_sum += int(t.get('cal').get('kal'))
            quanty_sum += int(t.get('cal').get('quanty'))
            tan_g_sum += int(t.get('cal').get('tan_g'))
            dan_g_sum += int(t.get('cal').get('dan_g'))
            ji_g_sum += int(t.get('cal').get('ji_g'))
            dang_g_sum += int(t.get('cal').get('dang_g'))
            na_mg_sum += int(t.get('cal').get('na_mg'))
            chol_mg_sum += int(t.get('cal').get('chol_mg'))
            pho_g_sum += int(t.get('cal').get('pho_g'))

        ret_list = [kal_sum, quanty_sum, tan_g_sum, dan_g_sum, ji_g_sum, \
                    dang_g_sum, na_mg_sum, chol_mg_sum, pho_g_sum, trans_g_sum]

        print(ret_list)


    def set_bmi(self, weight_kg, height_cm):
        # height_cm = int(self.user_info.get('height'))
        # weight_kg = int(self.user_info.get('weight'))
        if isinstance(weight_kg, str):
            weight_kg = int(weight_kg)
        if isinstance(height_cm, str):
            height_cm = int(height_cm)

        if (weight_kg <= 40) or (weight_kg > 600) or (50 > height_cm):
            # height error, "Over the 50cm", weight error, "Between 41 and 599"
           self.bmi["bmi"] = None 
        else:
            # BMI = 몸무게(kg) / 키(m) * 키(m)
            temp_bmi = round(weight_kg / (height_cm * height_cm) * 10000, 2)
            if temp_bmi < 18.5:
                ret = "저체중"
            elif temp_bmi >= 18.5 and temp_bmi < 24.9:
                ret = "정상체중"
            elif temp_bmi >= 25 and temp_bmi < 29.9:
                ret = "과체중"
            else:
                ret = "비만"
            # str(parse(time.strftime('%Y-%m-%d')))
            self.bmi["bmi"] = [temp_bmi, ret]
    

    def get_energy(self):    
       height_cm = int(self.user_info.get('height'))
       weight_kg = int(self.user_info.get('weight'))
       age = int(self.user_info.get('age'))
       gender = int(self.user_info.get('gender'))
       pa = int(self.user_info.get('pa'))
       
       if (age <= 12) or (age > 80):
           return 11 # age error, "Between 13 and 80"
       if (weight_kg <= 40) or (weight_kg > 600):
           return 12 # weight error, "Between 41 and 599"
       if (50 > height_cm):
           return 13 # height error, "Over the 50cm"

       if (height_cm and weight_kg and age and gender) is not None:
            if gender == 1:
                # 성인 남성
               energe = round(662 - (9.53 * age) + pa * ((15.91 * weight_kg) 
                              + (539.6 * height_cm) / 100))
            else:
                # 성인 여성
               energe = round(354 - (6.91 * age) + pa * ((9.36 * weight_kg) 
                              + (726 * height_cm) / 100))
            return energe
       else:
           return None # 필요한 정보 부족


    '''
    탄수화물 = 2,000 x 0.65 = 1,300kcal/4= 325g
    단백질 = 2,000 x 0.15 = 300kcal/4= 75g
    지방 = 2,000 x 0.2 = 400kcal/9= 44g
    (3대 열량영양소 섭취비율은 대상자에 따라 적절히 조절)
    '''
    def major_nutrients(self):
        energe = self.get_energy()
        tan_g = "%s" % str(round((energe * 0.65) / 4)) + "g"
        dan_g = "%s" % str(round((energe * 0.15) / 4)) + "g"
        ji_g =  "%s" % str(round((energe * 0.2) / 9)) + "g"
        result = {
            "tan_g": tan_g,
            "dan_g": dan_g,
            "ji_g": ji_g
        }
        return result


    def cal_clt(self, _type=None, _mins=None, _cals=0):
        height_cm = int(self.user_info.get('height'))
        weight_kg = int(self.user_info.get('weight'))
        age = int(self.user_info.get('age'))
        gender = int(self.user_info.get('gender'))
        
        if isinstance(_mins, str):
            p = re.compile('\d+')
            m = p.match(_mins)
            if m:
                _mins = int(m.group())
            else:
                return 14 # mins or cals error
        if _mins is None or not _mins: 
            return 14 # mins or cals error
        if (age <= 12) or (age > 80):
            return 11 # age error, "Between 13 and 80"
        if (weight_kg <= 40) or (weight_kg > 600):
            return 12 # weight error, "Between 41 and 599"
        if (50 > height_cm):
            return 13 # height error, "Over the 50cm"
        if (_mins > 0 and _cals > 0) or (_mins == 0 and _cals == 0):
	        return 14 # mins or cals error

        if gender == 1:
            # 성인 남성
            result = 66.5 + (13.75 * weight_kg) + (5.003 * height_cm) - (6.775 * age)
        else:
            # 성인 여성
            result = 655.1 + (9.563 * weight_kg) + (1.850 * height_cm) - (4.676 * age); 
        
        distance = min(self.exercise, key = lambda s: self.edit_distance(_type, s[0]))
        # [자전거, 걷기, 조깅, 수영, 야구, 볼링, 복싱, 축구, 스쿼시, 테니스, 배구, 코어, 필라테스, PT, 에어로빅, 헬스]
        if _mins > 0:
            value = round((result * distance[1]) * (_mins/1440))
        if _cals > 0:
            value = round((_cals / result / distance[1]) * (1440))
        
        return str(value) + "cal"


if __name__ == "__main__":
    start = time.time()
    fb = Firebase()
    fb.set_cal("참치김밥")
    print(fb.ret)

    # fb.set_bmi(75, 178) # input: weight_kg, height_cm
    # fb.user_info['height'] = 178 # input: height_cm
    # fb.user_info['weight'] = 75 # input: weight_cm
    # fb.user_info['age'] = 26 # input: age_number
    # fb.user_info['gender'] = 1 # input: gender_MaleorFemale
    # fb.user_info['pa'] = fb.m_pa_list[1] # input: 신체 활동 계수, 비활동적, 보통, 활동적, 매우 활동적
    # # fb.user_info['pa'] = fb.fm_pa_list[1]
    # print(fb.bmi)
    # print(str(fb.get_energy())+"kcal") 
    # print(fb.cal_clt(_type="자전거", _mins="60")) # input: _type, MinsorCals
    # print(fb.cal_clt(_type="자전거", _mins="반")) # input: _type, MinsorCals
    # print(fb.cal_clt(_type="자전거", _mins="60분")) # input: _type, MinsorCals
    # print(fb.major_nutrients())

    '''
    test -> e.g.

    {'2018-11-20 00:00:00': [23.67, '정상체중']}
    2568kcal
    566cal
    318cal
    {'탄수화물': '417g', '단백질': '96g', '지방': '57g'}
    '''
    print("--- %s seconds---" % (time.time() - start))
